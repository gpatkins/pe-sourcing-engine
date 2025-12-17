"""
PE Sourcing Engine v5.1 - FastAPI Dashboard
Main API application with JWT-based authentication and user management.
"""

from __future__ import annotations

import json
import os
import secrets
import csv
import io
from pathlib import Path
from typing import Any, Dict, List, Optional
from datetime import datetime, timedelta

import yaml
from dotenv import load_dotenv
from fastapi import (
    BackgroundTasks,
    Cookie,
    Depends,
    FastAPI,
    Form,
    HTTPException,
    Request,
    Response,
)
from fastapi.responses import RedirectResponse, JSONResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from starlette.status import HTTP_303_SEE_OTHER

from etl.utils.state_manager import get_current_state, request_stop
from etl.utils.db import get_db_connection, fetch_all_dict, fetch_one_dict, execute

# Authentication imports (v5.1)
from api.auth import (
    verify_password,
    get_password_hash,
    create_access_token,
    validate_email_address,
    validate_password_strength,
    COOKIE_NAME,
    ACCESS_TOKEN_EXPIRE_MINUTES,
)
from api.dependencies import (
    get_current_user,
    get_current_user_optional,
    get_current_active_user,
    require_admin,
    get_user_companies_filter,
    log_user_activity,
)
from api.models import ActivityType

# Pipeline Imports
try:
    from run_pipeline import run_discover, run_enrich, run_score
except ImportError:
    def run_discover(): pass
    def run_enrich(): pass
    def run_score(): pass

BASE_DIR = Path(__file__).resolve().parent.parent
CONFIG_PATH = BASE_DIR / "config" / "settings.yaml"
ENV_PATH = BASE_DIR / "config" / "secrets.env"
LOGS_DIR = BASE_DIR / "logs"
LOG_FILE = LOGS_DIR / "pipeline.log"

load_dotenv(ENV_PATH)

# Metabase URL configuration (v5.2)
METABASE_URL = os.getenv("METABASE_URL", "http://localhost:3000")

# --- FastAPI App ---
app = FastAPI(title="DealGenome - PE Sourcing Engine v5.1")
templates = Jinja2Templates(directory="api/templates")

# Add current_user to all template contexts
@app.middleware("http")
async def add_user_to_templates(request: Request, call_next):
    """Middleware to add current user to all template contexts."""
    response = await call_next(request)
    return response

# --- CSRF Protection ---
CSRF_SECRET = os.getenv("CSRF_SECRET", secrets.token_hex(32))

def generate_csrf_token() -> str:
    """Generate CSRF token for forms."""
    return secrets.token_urlsafe(32)

def verify_csrf_token(token: str, expected: str) -> bool:
    """Verify CSRF token matches."""
    import hmac
    return hmac.compare_digest(token, expected)

# --- SETTINGS HELPERS ---
def load_settings() -> Dict[str, Any]:
    if not CONFIG_PATH.exists(): return {}
    with CONFIG_PATH.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}

def save_settings(settings: Dict[str, Any]) -> None:
    with CONFIG_PATH.open("w", encoding="utf-8") as f:
        yaml.safe_dump(settings, f, sort_keys=False)

def get_discovery_queries() -> List[Dict[str, Any]]:
    settings = load_settings()
    discovery = settings.get("discovery", {})
    queries = discovery.get("queries", [])
    return queries

def update_discovery_queries(queries: List[Dict[str, Any]]) -> None:
    settings = load_settings()
    if "discovery" not in settings: settings["discovery"] = {}
    settings["discovery"]["queries"] = queries
    save_settings(settings)

# --- STATUS & STATS ---
def get_dashboard_stats(user_id: Optional[int] = None, is_admin: bool = False) -> dict:
    """Get dashboard statistics (all users see all companies)."""
    conn = get_db_connection()
    try:
        cur = conn.cursor()

        # All users see all companies (no user-based filtering)
        # Admin vs regular user distinction is for user management only, not data isolation
        
        # 1. Total Discovered
        cur.execute("SELECT COUNT(*) FROM companies;")
        total = cur.fetchone()[0]

        # 2. Fully Enriched (Completed)
        cur.execute("SELECT COUNT(*) FROM companies WHERE enrichment_status = 'complete';")
        enriched = cur.fetchone()[0]

        # 3. Pending Enrichment (Queue)
        cur.execute("SELECT COUNT(*) FROM companies WHERE enrichment_status IN ('pending', 'partial');")
        pending = cur.fetchone()[0]

        # 4. Scored Companies
        cur.execute("SELECT COUNT(*) FROM companies WHERE buyability_score IS NOT NULL;")
        scored = cur.fetchone()[0]

        # 5. Risk Alerts
        try:
            cur.execute("SELECT COUNT(*) FROM companies WHERE risk_flags LIKE 'ALERT%';")
            risks = cur.fetchone()[0]
        except:
            risks = 0

        return {
            "discovered_count": total,
            "enriched_count": enriched,
            "pending_count": pending,
            "scored_count": scored,
            "risk_count": risks
        }
    except Exception as e:
        print(f"Error getting stats: {e}")
        return {"discovered_count": 0, "enriched_count": 0, "pending_count": 0, "scored_count": 0, "risk_count": 0}
    finally:
        conn.close()

def get_user_stats(user_id: int) -> dict:
    """Get user-specific statistics for profile page."""
    sql = """
        SELECT
            COUNT(*) as total_companies,
            COUNT(CASE WHEN created_at > NOW() - INTERVAL '30 days' THEN 1 END) as companies_last_30_days,
            MAX(created_at) as last_company_added
        FROM companies
        WHERE user_id = %s
    """
    result = fetch_one_dict(sql, (user_id,))
    return result or {"total_companies": 0, "companies_last_30_days": 0, "last_company_added": None}

# ============================================================================
# AUTHENTICATION ROUTES (v5.1)
# ============================================================================

@app.get("/login")
async def login_page(request: Request, current_user: Optional[dict] = Depends(get_current_user_optional)):
    """Display login page."""
    # If already logged in, redirect to dashboard
    if current_user:
        return RedirectResponse(url="/dashboard", status_code=HTTP_303_SEE_OTHER)

    csrf_token = generate_csrf_token()
    return templates.TemplateResponse("login.html", {
        "request": request,
        "csrf_token": csrf_token,
        "current_user": None
    })

@app.post("/login")
async def login(
    request: Request,
    response: Response,
    email: str = Form(...),
    password: str = Form(...),
):
    """Process login and set auth cookie."""
    # Validate credentials
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT id, email, hashed_password, full_name, role, is_active FROM users WHERE email = %s", (email,))
    user = cur.fetchone()

    if not user:
        return templates.TemplateResponse("login.html", {
            "request": request,
            "csrf_token": generate_csrf_token(),
            "messages": [{"type": "error", "text": "Invalid email or password"}],
            "email": email,
            "current_user": None
        }, status_code=400)

    user_id, user_email, hashed_password, full_name, role, is_active = user

    # Verify password
    if not verify_password(password, hashed_password):
        return templates.TemplateResponse("login.html", {
            "request": request,
            "csrf_token": generate_csrf_token(),
            "messages": [{"type": "error", "text": "Invalid email or password"}],
            "email": email,
            "current_user": None
        }, status_code=400)

    # Check if active
    if not is_active:
        return templates.TemplateResponse("login.html", {
            "request": request,
            "csrf_token": generate_csrf_token(),
            "messages": [{"type": "error", "text": "Your account is inactive. Contact administrator."}],
            "email": email,
            "current_user": None
        }, status_code=403)

    # Update last login
    cur.execute("UPDATE users SET last_login = NOW() WHERE id = %s", (user_id,))
    conn.commit()

    # Log activity
    log_user_activity(conn, user_id, ActivityType.LOGIN, {"ip": request.client.host})

    conn.close()

    # Create JWT token
    access_token = create_access_token(
        data={"user_id": user_id, "email": user_email, "role": role}
    )

    # Set cookie and redirect
    redirect = RedirectResponse(url="/dashboard", status_code=HTTP_303_SEE_OTHER)
    redirect.set_cookie(
        key=COOKIE_NAME,
        value=f"Bearer {access_token}",
        max_age=ACCESS_TOKEN_EXPIRE_MINUTES * 60,
        httponly=True,
        samesite="lax"
    )

    return redirect

@app.get("/logout")
async def logout(response: Response):
    """Logout user by clearing auth cookie."""
    redirect = RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)
    redirect.delete_cookie(key=COOKIE_NAME)
    return redirect

@app.get("/register")
async def register_page(request: Request, admin: dict = Depends(require_admin)):
    """Display user registration page (admin only)."""
    csrf_token = generate_csrf_token()
    return templates.TemplateResponse("register.html", {
        "request": request,
        "csrf_token": csrf_token,
        "current_user": admin
    })

@app.post("/register")
async def register_user(
    request: Request,
    email: str = Form(...),
    password: str = Form(...),
    confirm_password: str = Form(...),
    full_name: str = Form(""),
    role: str = Form("user"),
    is_active: bool = Form(True),
    admin: dict = Depends(require_admin)
):
    """Create new user (admin only)."""
    # Validate email
    if not validate_email_address(email):
        return templates.TemplateResponse("register.html", {
            "request": request,
            "csrf_token": generate_csrf_token(),
            "messages": [{"type": "error", "text": "Invalid email address"}],
            "email": email,
            "full_name": full_name,
            "role": role,
            "current_user": admin
        }, status_code=400)

    # Validate passwords match
    if password != confirm_password:
        return templates.TemplateResponse("register.html", {
            "request": request,
            "csrf_token": generate_csrf_token(),
            "messages": [{"type": "error", "text": "Passwords do not match"}],
            "email": email,
            "full_name": full_name,
            "role": role,
            "current_user": admin
        }, status_code=400)

    # Validate password strength
    is_valid, error_msg = validate_password_strength(password)
    if not is_valid:
        return templates.TemplateResponse("register.html", {
            "request": request,
            "csrf_token": generate_csrf_token(),
            "messages": [{"type": "error", "text": error_msg}],
            "email": email,
            "full_name": full_name,
            "role": role,
            "current_user": admin
        }, status_code=400)

    # Check if user already exists
    existing = fetch_one_dict("SELECT id FROM users WHERE email = %s", (email,))
    if existing:
        return templates.TemplateResponse("register.html", {
            "request": request,
            "csrf_token": generate_csrf_token(),
            "messages": [{"type": "error", "text": "User with this email already exists"}],
            "email": email,
            "full_name": full_name,
            "role": role,
            "current_user": admin
        }, status_code=400)

    # Create user
    hashed_password = get_password_hash(password)
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO users (email, hashed_password, full_name, role, is_active)
        VALUES (%s, %s, %s, %s, %s)
        RETURNING id
        """,
        (email, hashed_password, full_name or None, role, is_active)
    )
    new_user_id = cur.fetchone()[0]
    conn.commit()

    # Log activity
    log_user_activity(conn, admin["user_id"], ActivityType.USER_CREATED, {"created_user_id": new_user_id, "email": email})

    conn.close()

    # Redirect with success message
    return RedirectResponse(url="/admin/users?success=User created successfully", status_code=HTTP_303_SEE_OTHER)

@app.get("/profile")
async def profile_page(request: Request, current_user: dict = Depends(get_current_active_user)):
    """Display user profile page."""
    # Get user details
    user = fetch_one_dict("SELECT * FROM users WHERE id = %s", (current_user["user_id"],))

    # Get user stats
    stats = get_user_stats(current_user["user_id"])

    # Parse query parameters for messages
    messages = []
    if request.query_params.get("success"):
        messages.append({"type": "success", "text": request.query_params.get("success")})
    if request.query_params.get("error"):
        messages.append({"type": "error", "text": request.query_params.get("error")})

    csrf_token = generate_csrf_token()
    return templates.TemplateResponse("profile.html", {
        "request": request,
        "user": user,
        "stats": stats,
        "csrf_token": csrf_token,
        "current_user": current_user,
        "messages": messages if messages else None
    })

@app.post("/profile/update-name")
async def update_profile_name(
    request: Request,
    full_name: str = Form(...),
    current_user: dict = Depends(get_current_active_user)
):
    """Update user's full name."""
    execute("UPDATE users SET full_name = %s WHERE id = %s", (full_name.strip() or None, current_user["user_id"]))

    return RedirectResponse(url="/profile?success=Name updated successfully", status_code=HTTP_303_SEE_OTHER)

@app.post("/profile/change-password")
async def change_password(
    request: Request,
    current_password: str = Form(...),
    new_password: str = Form(...),
    confirm_new_password: str = Form(...),
    current_user: dict = Depends(get_current_active_user)
):
    """Change user password."""
    # Verify current password
    user = fetch_one_dict("SELECT hashed_password FROM users WHERE id = %s", (current_user["user_id"],))

    if not verify_password(current_password, user["hashed_password"]):
        return RedirectResponse(url="/profile?error=Current password is incorrect", status_code=HTTP_303_SEE_OTHER)

    # Validate new passwords match
    if new_password != confirm_new_password:
        return RedirectResponse(url="/profile?error=New passwords do not match", status_code=HTTP_303_SEE_OTHER)

    # Validate password strength
    is_valid, error_msg = validate_password_strength(new_password)
    if not is_valid:
        return RedirectResponse(url=f"/profile?error={error_msg}", status_code=HTTP_303_SEE_OTHER)

    # Update password
    hashed_password = get_password_hash(new_password)
    execute("UPDATE users SET hashed_password = %s WHERE id = %s", (hashed_password, current_user["user_id"]))

    # Log activity
    conn = get_db_connection()
    log_user_activity(conn, current_user["user_id"], ActivityType.PASSWORD_CHANGED, {})
    conn.close()

    return RedirectResponse(url="/profile?success=Password changed successfully", status_code=HTTP_303_SEE_OTHER)

# ============================================================================
# ADMIN ROUTES (v5.1)
# ============================================================================

@app.get("/admin/users")
async def admin_users_page(request: Request, admin: dict = Depends(require_admin)):
    """Admin user management page."""
    # Get all users with company counts
    users = fetch_all_dict("""
    SELECT
        u.id, u.email, u.full_name, u.role, u.is_active, u.created_at, u.last_login,
        COUNT(c.id) as company_count
    FROM users u
    LEFT JOIN companies c ON (c.user_id = u.id OR (u.role = 'admin' AND c.user_id IS NULL))
    GROUP BY u.id, u.email, u.full_name, u.role, u.is_active, u.created_at, u.last_login
    ORDER BY u.created_at DESC
""")

    # Get recent activity
    recent_activity = fetch_all_dict("""
        SELECT
            ua.id, ua.user_id, ua.activity_type, ua.details, ua.created_at,
            u.email as user_email
        FROM user_activity ua
        JOIN users u ON u.id = ua.user_id
        ORDER BY ua.created_at DESC
        LIMIT 10
    """)

    # Parse query parameters for messages
    messages = []
    if request.query_params.get("success"):
        messages.append({"type": "success", "text": request.query_params.get("success")})
    if request.query_params.get("error"):
        messages.append({"type": "error", "text": request.query_params.get("error")})

    csrf_token = generate_csrf_token()
    return templates.TemplateResponse("admin_users.html", {
        "request": request,
        "users": users,
        "recent_activity": recent_activity,
        "csrf_token": csrf_token,
        "current_user": admin,
        "messages": messages if messages else None
    })

@app.post("/admin/clear-companies")
async def clear_companies_database(admin: dict = Depends(require_admin)):
    """Clear all companies from the database (admin only)."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Truncate companies table (cascades to related data if needed)
        cur.execute("TRUNCATE TABLE companies RESTART IDENTITY CASCADE;")
        conn.commit()
        
        # Log the activity
        log_user_activity(conn, admin["user_id"], ActivityType.DISCOVERY_RUN, {"action": "clear_companies_database"})
        
        cur.close()
        conn.close()
        
        return JSONResponse({
            "success": True,
            "message": "All companies cleared from database"
        })
    except Exception as e:
        return JSONResponse({
            "success": False,
            "message": f"Error clearing database: {str(e)}"
        }, status_code=500)

@app.post("/admin/users/toggle-status/{user_id}")
async def toggle_user_status(
    user_id: int,
    admin: dict = Depends(require_admin)
):
    """Toggle user active/inactive status."""
    # Prevent admin from deactivating themselves
    if user_id == admin["user_id"]:
        return RedirectResponse(url="/admin/users?error=Cannot deactivate your own account", status_code=HTTP_303_SEE_OTHER)

    # Toggle status
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("UPDATE users SET is_active = NOT is_active WHERE id = %s RETURNING is_active", (user_id,))
    new_status = cur.fetchone()[0]
    conn.commit()

    # Log activity
    log_user_activity(conn, admin["user_id"], ActivityType.USER_UPDATED, {"user_id": user_id, "action": "toggled_status", "new_status": new_status})

    conn.close()

    status_text = "activated" if new_status else "deactivated"
    return RedirectResponse(url=f"/admin/users?success=User {status_text} successfully", status_code=HTTP_303_SEE_OTHER)

@app.post("/admin/users/reset-password/{user_id}")
async def reset_user_password(
    user_id: int,
    admin: dict = Depends(require_admin)
):
    """Reset user password to a temporary password."""
    # Generate temporary password
    temp_password = secrets.token_urlsafe(12)
    hashed_password = get_password_hash(temp_password)

    # Update password
    execute("UPDATE users SET hashed_password = %s WHERE id = %s", (hashed_password, user_id))

    # Get user email
    user = fetch_one_dict("SELECT email FROM users WHERE id = %s", (user_id,))

    # Log activity
    conn = get_db_connection()
    log_user_activity(conn, admin["user_id"], ActivityType.USER_UPDATED, {"user_id": user_id, "action": "password_reset"})
    conn.close()

    # In production, you'd email this to the user. For now, show it to admin.
    return RedirectResponse(
        url=f"/admin/users?success=Password reset for {user['email']}. Temporary password: {temp_password}",
        status_code=HTTP_303_SEE_OTHER
    )

@app.post("/admin/users/delete/{user_id}")
async def delete_user(
    user_id: int,
    admin: dict = Depends(require_admin)
):
    """Delete a user."""
    # Prevent admin from deleting themselves
    if user_id == admin["user_id"]:
        return RedirectResponse(url="/admin/users?error=Cannot delete your own account", status_code=HTTP_303_SEE_OTHER)

    # Get user email for logging
    user = fetch_one_dict("SELECT email FROM users WHERE id = %s", (user_id,))

    # Delete user (cascade will handle user_activity)
    execute("DELETE FROM users WHERE id = %s", (user_id,))

    # Log activity
    conn = get_db_connection()
    log_user_activity(conn, admin["user_id"], ActivityType.USER_DELETED, {"deleted_user_id": user_id, "email": user["email"]})
    conn.close()

    return RedirectResponse(url="/admin/users?success=User deleted successfully", status_code=HTTP_303_SEE_OTHER)

@app.get("/admin/api-keys")
async def admin_api_keys_page(request: Request, admin: dict = Depends(require_admin)):
    """Admin API keys management page."""
    # Get all API credentials
    credentials = fetch_all_dict("""
        SELECT id, service_name, api_key, is_active, updated_at, updated_by
        FROM api_credentials
        ORDER BY service_name
    """)

    # Add masked keys for display
    for cred in credentials:
        if cred['api_key'] and len(cred['api_key']) > 8:
            cred['masked_key'] = '*' * (len(cred['api_key']) - 4) + cred['api_key'][-4:]
        else:
            cred['masked_key'] = '****'

    # Parse query parameters for messages
    messages = []
    if request.query_params.get("success"):
        messages.append({"type": "success", "text": request.query_params.get("success")})
    if request.query_params.get("error"):
        messages.append({"type": "error", "text": request.query_params.get("error")})

    csrf_token = generate_csrf_token()
    return templates.TemplateResponse("admin_api_keys.html", {
        "request": request,
        "credentials": credentials,
        "csrf_token": csrf_token,
        "current_user": admin,
        "messages": messages if messages else None
    })

@app.post("/admin/api-keys/update/{credential_id}")
async def update_api_key(
    credential_id: int,
    api_key: str = Form(...),
    admin: dict = Depends(require_admin)
):
    """Update an API key."""
    if not api_key or not api_key.strip():
        return RedirectResponse(url="/admin/api-keys?error=API key cannot be empty", status_code=HTTP_303_SEE_OTHER)

    # Update the API key
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "UPDATE api_credentials SET api_key = %s, updated_at = NOW(), updated_by = %s WHERE id = %s RETURNING service_name",
        (api_key.strip(), admin["user_id"], credential_id)
    )
    result = cur.fetchone()

    if not result:
        conn.close()
        return RedirectResponse(url="/admin/api-keys?error=API credential not found", status_code=HTTP_303_SEE_OTHER)

    service_name = result[0]
    conn.commit()

    # Log activity
    log_user_activity(conn, admin["user_id"], ActivityType.API_KEY_UPDATE, {"credential_id": credential_id, "service": service_name})

    conn.close()

    return RedirectResponse(url=f"/admin/api-keys?success={service_name} API key updated successfully", status_code=HTTP_303_SEE_OTHER)

@app.post("/admin/api-keys/toggle/{credential_id}")
async def toggle_api_key_status(
    credential_id: int,
    admin: dict = Depends(require_admin)
):
    """Toggle API key active/inactive status."""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "UPDATE api_credentials SET is_active = NOT is_active WHERE id = %s RETURNING is_active, service_name",
        (credential_id,)
    )
    result = cur.fetchone()

    if not result:
        conn.close()
        return RedirectResponse(url="/admin/api-keys?error=API credential not found", status_code=HTTP_303_SEE_OTHER)

    new_status, service_name = result
    conn.commit()

    # Log activity
    log_user_activity(conn, admin["user_id"], ActivityType.API_KEY_UPDATE, {
        "credential_id": credential_id,
        "service": service_name,
        "action": "toggled_status",
        "new_status": new_status
    })

    conn.close()

    status_text = "enabled" if new_status else "disabled"
    return RedirectResponse(url=f"/admin/api-keys?success={service_name} API key {status_text}", status_code=HTTP_303_SEE_OTHER)

# ============================================================================
# DASHBOARD ROUTES (Protected)
# ============================================================================

@app.get("/")
async def root(current_user: Optional[dict] = Depends(get_current_user_optional)):
    """
    Redirect root to appropriate page based on authentication status.
    - Authenticated users → /dashboard
    - Unauthenticated users → /login
    """
    if current_user:
        return RedirectResponse(url="/dashboard")
    else:
        return RedirectResponse(url="/login")

@app.get("/dashboard")
async def dashboard(request: Request, current_user: dict = Depends(get_current_active_user)):
    """Main dashboard page."""
    is_admin = current_user.get("role") == "admin"
    user_id = None if is_admin else current_user["user_id"]
    stats = get_dashboard_stats(user_id, is_admin)
    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "stats": stats,
        "current_user": current_user,
        "metabase_url": METABASE_URL
    })

@app.get("/companies")
async def companies_page(request: Request, current_user: dict = Depends(get_current_active_user)):
    """Companies export page."""
    is_admin = current_user.get("role") == "admin"
    user_id = None if is_admin else current_user["user_id"]

    # Get stats for the page
    where_clause = "" if is_admin else "WHERE user_id = %s"
    params = () if is_admin else (user_id,)

    conn = get_db_connection()
    cur = conn.cursor()

    # Total companies
    cur.execute(f"SELECT COUNT(*) FROM companies {where_clause}", params)
    total_companies = cur.fetchone()[0]

    # Enriched companies
    cur.execute(f"SELECT COUNT(*) FROM companies {where_clause} {'AND' if where_clause else 'WHERE'} enrichment_status = 'complete'", params)
    enriched_companies = cur.fetchone()[0]

    # Scored companies
    cur.execute(f"SELECT COUNT(*) FROM companies {where_clause} {'AND' if where_clause else 'WHERE'} buyability_score IS NOT NULL", params)
    scored_companies = cur.fetchone()[0]

    conn.close()

    stats = {
        "total_companies": total_companies,
        "enriched_companies": enriched_companies,
        "scored_companies": scored_companies
    }

    return templates.TemplateResponse("companies.html", {
        "request": request,
        "stats": stats,
        "current_user": current_user,
        "metabase_url": METABASE_URL
    })

@app.post("/companies/export/csv")
async def export_companies_csv(request: Request, current_user: dict = Depends(get_current_active_user)):
    """Export companies as CSV."""
    is_admin = current_user.get("role") == "admin"
    user_id = None if is_admin else current_user["user_id"]

    # Get filter from query params
    filter_type = request.query_params.get("filter", "all")

    # Build query based on filter
    where_clauses = []
    params = []

    if not is_admin:
        where_clauses.append("user_id = %s")
        params.append(user_id)

    if filter_type == "top_scored":
        where_clauses.append("buyability_score >= 60")
    elif filter_type == "family_owned":
        where_clauses.append("is_family_owned = TRUE")

    where_clause = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""

    sql = f"""
        SELECT
            name, url, phone, address, city, state, zip,
            industry_tag, customer_type, revenue_estimate, buyability_score,
            is_family_owned, is_franchise, owner_name, owner_phone,
            linkedin_company_url, owner_linkedin_url, founder_email,
            risk_flags, legal_name, naics_code, naics_description,
            website_tech_stack, google_rating, google_reviews,
            created_at, last_enriched_at
        FROM companies
        {where_clause}
        ORDER BY buyability_score DESC NULLS LAST, created_at DESC
    """

    companies = fetch_all_dict(sql, tuple(params) if params else None)

    # Create CSV in memory
    output = io.StringIO()
    if companies:
        fieldnames = companies[0].keys()
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        for company in companies:
            # Convert datetime objects to strings for CSV
            row = {}
            for key, value in company.items():
                if hasattr(value, 'strftime'):
                    row[key] = value.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(value, (dict, list)):
                    row[key] = str(value)
                else:
                    row[key] = value
            writer.writerow(row)

    output.seek(0)

    # Log export activity
    conn = get_db_connection()
    log_user_activity(conn, current_user["user_id"], ActivityType.EXPORT_DATA, {
        "format": "csv",
        "filter": filter_type,
        "count": len(companies)
    })
    conn.close()

    filename = f"companies_{filter_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.post("/companies/export/excel")
async def export_companies_excel(request: Request, current_user: dict = Depends(get_current_active_user)):
    """Export companies as Excel."""
    # Check if openpyxl is available
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel export requires openpyxl. Please install it: pip install openpyxl")

    is_admin = current_user.get("role") == "admin"
    user_id = None if is_admin else current_user["user_id"]

    # Get filter from query params
    filter_type = request.query_params.get("filter", "all")

    # Build query based on filter
    where_clauses = []
    params = []

    if not is_admin:
        where_clauses.append("user_id = %s")
        params.append(user_id)

    if filter_type == "top_scored":
        where_clauses.append("buyability_score >= 60")
    elif filter_type == "family_owned":
        where_clauses.append("is_family_owned = TRUE")

    where_clause = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""

    sql = f"""
        SELECT
            name, url, phone, address, city, state, zip,
            industry_tag, customer_type, revenue_estimate, buyability_score,
            is_family_owned, is_franchise, owner_name, owner_phone,
            linkedin_company_url, owner_linkedin_url, founder_email,
            risk_flags, legal_name, naics_code, naics_description,
            website_tech_stack, google_rating, google_reviews,
            created_at, last_enriched_at
        FROM companies
        {where_clause}
        ORDER BY buyability_score DESC NULLS LAST, created_at DESC
    """

    companies = fetch_all_dict(sql, tuple(params) if params else None)

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Companies"

    # Header styling
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    if companies:
        # Write headers
        headers = list(companies[0].keys())
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Write data
        for row_num, company in enumerate(companies, 2):
            for col_num, header in enumerate(headers, 1):
                value = company[header]
                # Convert datetime objects to strings
                if hasattr(value, 'strftime'):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                # Convert JSONB/dict to string
                elif isinstance(value, (dict, list)):
                    value = str(value)
                ws.cell(row=row_num, column=col_num, value=value)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Log export activity
    conn = get_db_connection()
    log_user_activity(conn, current_user["user_id"], ActivityType.EXPORT_DATA, {
        "format": "excel",
        "filter": filter_type,
        "count": len(companies)
    })
    conn.close()

    filename = f"companies_{filter_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        io.BytesIO(output.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/discovery")
async def discovery_page(request: Request, current_user: dict = Depends(get_current_active_user)):
    """Discovery queries management page."""
    queries = get_discovery_queries()
    return templates.TemplateResponse("discovery_queries.html", {
        "request": request,
        "queries": list(enumerate(queries)),
        "current_user": current_user
    })

@app.get("/logs")
async def view_logs(request: Request, current_user: dict = Depends(get_current_active_user)):
    """View pipeline logs."""
    lines = []
    if LOG_FILE.exists():
        with LOG_FILE.open("r", encoding="utf-8") as f:
            lines = [line.rstrip() for line in f.readlines()[-200:]]
    return templates.TemplateResponse("logs.html", {
        "request": request,
        "lines": lines,
        "current_user": current_user
    })

# ============================================================================
# DISCOVERY QUERY ROUTES
# ============================================================================

@app.post("/discovery-queries/add")
async def add_discovery_query(
    label: str = Form(...),
    text_query: str = Form(...),
    region_code: str = Form("US"),
    limit: int = Form(20),
    current_user: dict = Depends(get_current_active_user)
):
    """Add new discovery query."""
    queries = get_discovery_queries()
    queries.append({
        "label": label.strip() or text_query[:40],
        "text_query": text_query.strip(),
        "region_code": region_code.strip() or "US",
        "limit": limit
    })
    update_discovery_queries(queries)
    return RedirectResponse(url="/discovery", status_code=HTTP_303_SEE_OTHER)

@app.post("/discovery-queries/delete/{index}")
async def delete_discovery_query(
    index: int,
    current_user: dict = Depends(get_current_active_user)
):
    """Delete discovery query."""
    queries = get_discovery_queries()
    if 0 <= index < len(queries):
        queries.pop(index)
        update_discovery_queries(queries)
    return RedirectResponse(url="/discovery", status_code=HTTP_303_SEE_OTHER)

@app.post("/discovery-queries/generate")
async def generate_queries(
    base_query: str = Form(...),
    limit: int = Form(20),
    current_user: dict = Depends(get_current_active_user)
):
    """Generate discovery queries for all active cities from database."""
    queries = get_discovery_queries()
    base = base_query.strip()
    
    # Get active locations from database
    active_locations = fetch_all_dict("""
        SELECT city, state 
        FROM scale_generator_config 
        WHERE is_active = true 
        ORDER BY state, city
    """)
    
    for loc in active_locations:
        city = loc['city']
        state = loc['state']
        new_text = f"{base} in {city}, {state}"
        label = f"gen_{state}_{city}"
        queries.append({
            "label": label,
            "text_query": new_text,
            "region_code": "US",
            "limit": limit
        })
    
    update_discovery_queries(queries)
    return RedirectResponse(url="/discovery", status_code=HTTP_303_SEE_OTHER)

@app.post("/discovery-queries/delete-all")
async def delete_all_queries(current_user: dict = Depends(get_current_active_user)):
    """Delete all discovery queries."""
    update_discovery_queries([])
    return RedirectResponse(url="/discovery", status_code=HTTP_303_SEE_OTHER)

# ============================================================================
# SCALE GENERATOR CONFIG API (v5.2)
# ============================================================================

@app.get("/api/scale-generator/locations")
async def get_scale_generator_locations(current_user: dict = Depends(get_current_active_user)):
    """Get all scale generator locations."""
    try:
        locations = fetch_all_dict("""
            SELECT id, city, state, is_active, created_at
            FROM scale_generator_config
            ORDER BY state, city
        """)
        
        # Convert datetime to string for JSON serialization
        for loc in locations:
            if loc.get('created_at'):
                loc['created_at'] = loc['created_at'].isoformat()
        
        return JSONResponse({"success": True, "locations": locations})
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"ERROR in get_scale_generator_locations: {error_details}")
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

@app.post("/api/scale-generator/locations")
async def add_scale_generator_location(
    request: Request,
    current_user: dict = Depends(get_current_active_user)
):
    """Add a new scale generator location."""
    try:
        data = await request.json()
        city = data.get("city", "").strip()
        state = data.get("state", "").strip().upper()
        
        if not city or not state:
            return JSONResponse({"success": False, "message": "City and state are required"}, status_code=400)
        
        if len(state) != 2:
            return JSONResponse({"success": False, "message": "State must be 2-letter code"}, status_code=400)
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        cur.execute("""
            INSERT INTO scale_generator_config (city, state, is_active)
            VALUES (%s, %s, true)
            RETURNING id
        """, (city, state))
        
        location_id = cur.fetchone()[0]
        conn.commit()
        cur.close()
        conn.close()
        
        return JSONResponse({"success": True, "message": "Location added", "id": location_id})
    except Exception as e:
        if "unique constraint" in str(e).lower():
            return JSONResponse({"success": False, "message": "Location already exists"}, status_code=400)
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)


@app.put("/api/scale-generator/locations/{location_id}")
async def update_scale_generator_location(
    location_id: int,
    request: Request,
    current_user: dict = Depends(get_current_active_user)
):
    """Update a scale generator location."""
    try:
        data = await request.json()
        is_active = data.get("is_active")
        
        if is_active is None:
            return JSONResponse({"success": False, "message": "is_active is required"}, status_code=400)
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        cur.execute("""
            UPDATE scale_generator_config
            SET is_active = %s, updated_at = NOW()
            WHERE id = %s
        """, (is_active, location_id))
        
        conn.commit()
        cur.close()
        conn.close()
        
        return JSONResponse({"success": True, "message": "Location updated"})
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)


@app.delete("/api/scale-generator/locations/{location_id}")
async def delete_scale_generator_location(
    location_id: int,
    current_user: dict = Depends(get_current_active_user)
):
    """Delete a scale generator location."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        
        cur.execute("DELETE FROM scale_generator_config WHERE id = %s", (location_id,))
        
        conn.commit()
        cur.close()
        conn.close()
        
        return JSONResponse({"success": True, "message": "Location deleted"})
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

# ============================================================================
# PIPELINE EXECUTION ROUTES
# ============================================================================

def _run_step(step: str, user_id: int):
    """Background task to run pipeline step."""
    if step == "discover": 
        run_discover(user_id)
    elif step == "enrich": 
        run_enrich()
    elif step == "score": 
        run_score()
    elif step == "full":
        run_discover(user_id)
        run_enrich()
        run_score()

@app.post("/run/{step}")
async def run_step(
    step: str,
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_active_user)
):
    """Run pipeline step in background."""
    if step in {"discover", "enrich", "score", "full"}:
        background_tasks.add_task(_run_step, step, current_user["user_id"])

        # Log activity
        conn = get_db_connection()
        log_user_activity(conn, current_user["user_id"], ActivityType.DISCOVERY_RUN if step == "discover" else ActivityType.ENRICHMENT_RUN, {"step": step})
        conn.close()

    return RedirectResponse(url="/dashboard", status_code=HTTP_303_SEE_OTHER)
    
# ============================================================================
# API ROUTES
# ============================================================================

@app.get("/api/status")
async def get_pipeline_status(current_user: dict = Depends(get_current_active_user)):
    """Get current pipeline status."""
    current_job = get_current_state()
    is_admin = current_user.get("role") == "admin"
    user_id = None if is_admin else current_user["user_id"]
    stats = get_dashboard_stats(user_id, is_admin)

    return JSONResponse({
        "status": current_job,
        "stats": stats
    })

@app.post("/api/stop")
async def stop_pipeline(current_user: dict = Depends(get_current_active_user)):
    """Stop running pipeline."""
    request_stop()
    return JSONResponse({"status": "stopping"})

@app.get("/api/logs/stream")
async def get_logs_stream(current_user: dict = Depends(get_current_active_user)):
    """Stream recent log lines."""
    if not LOG_FILE.exists():
        return JSONResponse({"lines": ["Waiting for logs..."]})
    with LOG_FILE.open("r", encoding="utf-8") as f:
        lines = f.readlines()
        last_lines = [line.rstrip() for line in lines[-50:]]
    return JSONResponse({"lines": last_lines})

@app.get("/api/pipeline/progress")
async def get_pipeline_progress():
    """Get current pipeline progress for UI."""
    try:
        progress_file = Path(__file__).resolve().parent.parent / "pipeline_progress.json"
        if progress_file.exists():
            with open(progress_file, "r") as f:
                return json.load(f)
        else:
            return {
                "stage": "idle",
                "status": "idle",
                "current": 0,
                "total": 0,
                "message": "System idle",
                "updated_at": datetime.utcnow().isoformat()
            }
    except Exception as e:
        logger.error(f"Failed to read progress: {e}")
        return {
            "stage": "error",
            "status": "error",
            "current": 0,
            "total": 0,
            "message": "Failed to read progress",
            "updated_at": datetime.utcnow().isoformat()
        }


@app.get("/api/pipeline/logs/download")
async def download_pipeline_logs(current_user: dict = Depends(get_current_active_user)):
    """Download pipeline log file."""
    from fastapi.responses import FileResponse

    log_file = Path(__file__).resolve().parent.parent / "logs" / "pipeline.log"
    if not log_file.exists():
        raise HTTPException(status_code=404, detail="Log file not found")

    return FileResponse(
        log_file,
        filename=f"pipeline_log_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.log",
        media_type="text/plain"
    )
